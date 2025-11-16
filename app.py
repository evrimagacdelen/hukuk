import streamlit as st
import pickle
import numpy as np
import os
import pandas as pd
import google.generativeai as genai
import re
import matplotlib
matplotlib.use('Agg') # Streamlit Cloud üzerinde uyumluluk için
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import io

# ==============================================================================
# BÖLÜM 1: TAHMİN MODELİ İÇİN GEREKLİ SINIF VE FONKSİYONLAR
# ==============================================================================

# Gerekli kütüphaneleri ve temel sınıfları import ediyoruz.
from sklearn.base import BaseEstimator, ClassifierMixin, clone
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.dummy import DummyClassifier

# CustomLawClassifier Sınıf Tanımı (Unpickling için gerekli)
class CustomLawClassifier(BaseEstimator, ClassifierMixin):
    def __init__(self, base_estimator):
        self.base_estimator = base_estimator
        self.models = []
    def fit(self, X, Y):
        self.models = []
        for i in range(Y.shape[1]):
            y_subset = Y[:, i]
            unique_classes = np.unique(y_subset)
            if len(unique_classes) < 2:
                model = DummyClassifier(strategy="constant", constant=unique_classes[0])
            else:
                model = clone(self.base_estimator)
            model.fit(X, y_subset)
            self.models.append(model)
        return self
    def predict(self, X):
        return np.array([model.predict(X) for model in self.models]).T

# ==============================================================================
# BÖLÜM 2: EXCEL RAPORLAMA İÇİN GEREKLİ FONKSİYONLAR
# ==============================================================================

def cerrahi_analiz_tek_satir(metin):
    """Sorumlu unvanlarını metinden çıkaran fonksiyon."""
    BILINEN_UNVANLAR = sorted(['Harcama Yetkilisi', 'Gerçekleştirme Görevlisi', 'Muhasebe Yetkilisi', 'Üst Yönetici', 'Akademik Teşvik Komisyonu', 'Üniversite Yönetim Kurulu', 'Döner Sermaye Yürütme Kurulu', 'Fakülte Yönetim Kurulu', 'İtiraz Komisyonu', 'Birim Komisyon', 'Jüri', 'Üniversite Senatosu', 'Personel Daire Başkanı', 'Strateji Geliştirme Daire Başkanı', 'İdari ve Mali İşler Daire Başkanı', 'Sağlık Kültür ve Spor Daire Başkanı', 'Döner Sermaye İşletme Müdürü', 'Hastane Başmüdürü', 'Hukuk Müşaviri', 'Fakülte Sekreteri', 'Enstitü Sekreteri', 'Yüksekokul Sekreteri', 'Rektör Yardımcısı', 'Dekan Yardımcısı', 'Başhekim Yardımcısı', 'Müdür Yardımcısı', 'Yüksekokul Müdürü', 'Enstitü Müdürü', 'Merkez Müdürü', 'Şube Müdürü', 'Hastane Müdürü', 'Daire Başkanı', 'Rektör', 'Dekan', 'Başhekim', 'Genel Sekreter', 'Müdür', 'Memur', 'Şef', 'Tekniker', 'Sayman', 'Bilgisayar İşletmeni', 'Öğretim Üyesi', 'Başkan'], key=len, reverse=True)
    AKADEMIK_DESENLER = {'Prof. Dr.': r'prof\s*\.\s*dr', 'Doç. Dr.': r'doç\s*\.\s*dr', 'Yrd. Doç. Dr.': r'yrd\s*\.\s*doç\s*\.\s*dr', 'Dr. Öğr. Üyesi': r'dr\s*\.\s*öğr\s*\.\s*üyesi', 'Öğr. Gör.': r'öğr\s*\.\s*gör', 'Dr.': r'\bdr\b'}
    NORM_MAP = {'hy': 'Harcama Yetkilisi', 'gg': 'Gerçekleştirme Görevlisi', 'dekan v.': 'Dekan Vekili', 'dekan v': 'Dekan Vekili', 'rektör yrd.': 'Rektör Yardımcısı', 'rektör yrd': 'Rektör Yardımcısı', 'müdür v.': 'Müdür Vekili', 'müdür v': 'Müdür Vekili', 'müdür yrd.': 'Müdür Yardımcısı', 'müdür yrd': 'Müdür Yardımcısı', 'fakülte sekreter v.': 'Fakülte Sekreteri Vekili', 'fakülte sekreter v': 'Fakülte Sekreteri Vekili', 'fakülte sekreterv': 'Fakülte Sekreteri Vekili', 'fakül. sekr. vekili': 'Fakülte Sekreteri Vekili', 'yüksekokul sekreter v.': 'Yüksekokul Sekreteri Vekili', 'yüksekokul sekreter v': 'Yüksekokul Sekreteri Vekili', 'yüksekokul sek. v': 'Yüksekokul Sekreteri Vekili', 'genel sekreter v.': 'Genel Sekreter Vekili', 'genel sekreter v': 'Genel Sekreter Vekili', 'döner ser. işl. md. v.': 'Döner Sermaye İşletme Müdürü Vekili', 'işletme müd. v.': 'İşletme Müdürü Vekili', 'hastane md. yrd': 'Hastane Müdür Yardımcısı', 'has. baş müd.': 'Hastane Başmüdürü', 'üyk': 'Üniversite Yönetim Kurulu', 'dsyk': 'Döner Sermaye Yürütme Kurulu'}
    if not isinstance(metin, str) or metin in ["YOK", "Kaynakta Yok"]: return []
    anlamsız_ifadeler = ['zararın', 'tahsil edildiği', 'ilişik kalmadı', 'kastedilmektedir', 'implicit', 'münferiden sorumlu']
    if any(ifade in metin.lower() for ifade in anlamsız_ifadeler): return []
    roller_bu_satirda, kalan_metin = set(), metin
    for standart_ad, desen in AKADEMIK_DESENLER.items():
        if re.search(desen, kalan_metin, re.IGNORECASE):
            roller_bu_satirda.add(standart_ad); kalan_metin = re.sub(desen, '', kalan_metin, flags=re.IGNORECASE)
    for unvan in BILINEN_UNVANLAR:
        if unvan.lower() in kalan_metin.lower():
            rol = unvan
            if any(k in unvan for k in ['Kurulu', 'Komisyonu', 'Senatosu', 'Jüri']): rol += ' Üyesi'
            roller_bu_satirda.add(rol); kalan_metin = re.sub(re.escape(unvan), '', kalan_metin, flags=re.IGNORECASE)
    potansiyel_roller = re.split(r'[,/()]|\s+ve\s+|\s+ile\s+', kalan_metin)
    for parca in potansiyel_roller:
        temiz_parca = parca.strip().lower()
        if temiz_parca in NORM_MAP: roller_bu_satirda.add(NORM_MAP[temiz_parca])
        elif 'vekili' in temiz_parca or temiz_parca.endswith((' v', ' v.')):
            if 'dekan' in temiz_parca: roller_bu_satirda.add('Dekan Vekili')
            elif 'rektör' in temiz_parca: roller_bu_satirda.add('Rektör Vekili')
    return list(roller_bu_satirda)

def create_pie_chart(data, title, filename):
    """Pasta grafiği oluşturan yardımcı fonksiyon. Veri boşsa dosya oluşturmaz."""
    if data.empty:
        st.warning(f"'{title}' için veri bulunamadığından grafik oluşturulmadı.")
        return False
    
    plt.figure(figsize=(8, 6))
    plt.pie(data, labels=data.index, autopct='%1.1f%%', startangle=140,
            wedgeprops={'edgecolor': 'white'}, textprops={'fontsize': 12})
    plt.title(title, fontsize=16, pad=20, weight='bold')
    plt.axis('equal')
    plt.savefig(filename, bbox_inches='tight', format='png')
    plt.close()
    return True

def generate_excel_report(script_dir):
    """Excel'den veri okuyup analiz ederek rapor oluşturan ana fonksiyon."""
    try:
        dosya_adi = os.path.join(script_dir, "sorumlu.xlsx")
        df = pd.read_excel(dosya_adi, sheet_name='VERİ-2-EMİR', header=0, dtype=str).fillna('')
        st.info(f"'{os.path.basename(dosya_adi)}' dosyasından {len(df)} satır veri bulundu.")

        sutun_map = {'Kararların Niteliği': 'Karar_Turu', 'Kamu Zararı Var mı?': 'Kamu_Zarari_Durumu', 'Kamu Zararının Sorumlusu Kim?': 'Sorumlular_Metni', 'Kararda Hangi Kanunlara ve Kanun Maddelerine Atıf Yapılmıştır?': 'Kanun_Maddeleri', 'Kararın Konusu Nedir?': 'Karar_Konusu', 'Azınlık Oyu': 'Azinlik_Oyu', 'Daire ilk kararında ısrar etmiş mi?': 'Israr_Durumu'}
        df.rename(columns=sutun_map, inplace=True)

        df['Azinlik_Oyu'] = df['Azinlik_Oyu'].str.strip()
        df['Israr_Durumu'] = df['Israr_Durumu'].str.strip()
        df['_KamuZarariVar'] = df['Kamu_Zarari_Durumu'].str.contains('Var|Zarar Oluştu', case=False, na=False)
        df['_AzinlikOyuVar'] = df['Azinlik_Oyu'].str.upper() == 'VAR'
        df['_IsrarVar'] = df['Israr_Durumu'] != ''
        st.info("Veri temizlendi ve yardımcı analiz sütunları oluşturuldu.")

        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            
            # SEKME 1: GENEL ÖZETLER
            st.info("Sekme 1: Genel Özetler ve Grafikler oluşturuluyor...")
            karar_turu_sayim = df['Karar_Turu'].value_counts()
            karsi_oy_sayim = df['Azinlik_Oyu'].value_counts()
            kamu_zarari_sayim = df['_KamuZarariVar'].value_counts().rename({True: 'Kamu Zararı Var', False: 'Kamu Zararı Yok'})
            israr_sayim = df[df['_IsrarVar']]['Israr_Durumu'].value_counts()
            
            karar_turu_sayim.to_excel(writer, sheet_name='Genel_Ozetler', header=['Sayı'], startrow=1, startcol=0); writer.sheets['Genel_Ozetler'].cell(1, 1).value = '1. Karar Türü Dağılımı'
            karsi_oy_sayim.to_excel(writer, sheet_name='Genel_Ozetler', header=['Sayı'], startrow=1, startcol=4); writer.sheets['Genel_Ozetler'].cell(1, 5).value = '2. Karşı Oy Dağılımı'
            kamu_zarari_sayim.to_excel(writer, sheet_name='Genel_Ozetler', header=['Sayı'], startrow=1, startcol=8); writer.sheets['Genel_Ozetler'].cell(1, 9).value = '3. Kamu Zararı Dağılımı'
            israr_sayim.to_excel(writer, sheet_name='Genel_Ozetler', header=['Sayı'], startrow=1, startcol=12); writer.sheets['Genel_Ozetler'].cell(1, 13).value = '7. Israr Kararı Dağılımı'
            
            ws = writer.sheets['Genel_Ozetler']
            chart_files = []
            try:
                if create_pie_chart(karar_turu_sayim, 'Karar Türü Dağılımı', 'chart1.png'):
                    ws.add_image(Image('chart1.png'), 'A6'); chart_files.append('chart1.png')
                if create_pie_chart(karsi_oy_sayim, 'Karşı Oy Dağılımı', 'chart2.png'):
                    ws.add_image(Image('chart2.png'), 'E6'); chart_files.append('chart2.png')
                if create_pie_chart(kamu_zarari_sayim, 'Kamu Zararı Dağılımı', 'chart3.png'):
                    ws.add_image(Image('chart3.png'), 'I6'); chart_files.append('chart3.png')
                if create_pie_chart(israr_sayim, 'Israr Kararı Dağılımı', 'chart4.png'):
                    ws.add_image(Image('chart4.png'), 'M6'); chart_files.append('chart4.png')
            finally:
                for f in chart_files:
                    if os.path.exists(f): os.remove(f)

            ct_kararturu_karsioy = pd.crosstab(df['Karar_Turu'], df['_AzinlikOyuVar']).rename(columns={True:'Var', False:'Yok'})
            ct_kamuzarari_karsioy = pd.crosstab(df['_KamuZarariVar'], df['_AzinlikOyuVar']).rename(index={True:'KZ Var', False:'KZ Yok'}, columns={True:'Var', False:'Yok'})
            ct_kararturu_karsioy.to_excel(writer, sheet_name='Genel_Ozetler', startrow=30, startcol=0); writer.sheets['Genel_Ozetler'].cell(30, 1).value = 'Karar Türü vs Karşı Oy'
            ct_kamuzarari_karsioy.to_excel(writer, sheet_name='Genel_Ozetler', startrow=30, startcol=5); writer.sheets['Genel_Ozetler'].cell(30, 6).value = 'Kamu Zararı vs Karşı Oy'
            
            # SEKME 2: UNVAN & KAMU ZARARI ANALİZİ
            st.info("Sekme 2: Unvan & Kamu Zararı Analizi oluşturuluyor...")
            analiz_listesi = []
            for _, satir in df.dropna(subset=['Sorumlular_Metni']).iterrows():
                unvanlar = cerrahi_analiz_tek_satir(satir['Sorumlular_Metni'])
                for unvan in unvanlar:
                    analiz_listesi.append({'Unvan': unvan, 'Zarar_Durumu': satir['_KamuZarariVar']})
            if analiz_listesi:
                ozet_tablo_unvan = pd.DataFrame(analiz_listesi).groupby('Unvan')['Zarar_Durumu'].value_counts().unstack(fill_value=0).rename(columns={True:'Kamu Zararı Var', False:'Kamu Zararı Yok'})
                if 'Kamu Zararı Var' not in ozet_tablo_unvan: ozet_tablo_unvan['Kamu Zararı Var'] = 0
                if 'Kamu Zararı Yok' not in ozet_tablo_unvan: ozet_tablo_unvan['Kamu Zararı Yok'] = 0
                ozet_tablo_unvan['Toplam'] = ozet_tablo_unvan.sum(axis=1)
                ozet_tablo_unvan['KZ Oranı %'] = ((ozet_tablo_unvan['Kamu Zararı Var'] / ozet_tablo_unvan['Toplam']) * 100).round(1)
                ozet_tablo_unvan.sort_values(by='Toplam', ascending=False).to_excel(writer, sheet_name='Unvan_Kamu_Zarari_Analizi')

            # SEKME 3: KARŞI OY DETAYLARI
            st.info("Sekme 3: Karşı Oy Detayları oluşturuluyor...")
            df_karsi_oy = df[df['_AzinlikOyuVar']].copy()
            if not df_karsi_oy.empty:
                karsi_oy_konu = df_karsi_oy['Karar_Konusu'].value_counts().reset_index().rename(columns={'index': 'Konu', 'Karar_Konusu': 'Sayı'})
                karsi_oy_kanun = df_karsi_oy['Kanun_Maddeleri'].value_counts().reset_index().rename(columns={'index': 'Kanun Maddesi', 'Kanun_Maddeleri': 'Sayı'})
                karsi_oy_konu.to_excel(writer, sheet_name='Karsi_Oy_Detaylari', startrow=1, startcol=0, index=False); writer.sheets['Karsi_Oy_Detaylari'].cell(1, 1).value = 'Karşı Oy Konuları'
                karsi_oy_kanun.to_excel(writer, sheet_name='Karsi_Oy_Detaylari', startrow=1, startcol=3, index=False); writer.sheets['Karsi_Oy_Detaylari'].cell(1, 4).value = 'Karşı Oy Kanun Maddeleri'

            # SEKME 4: KAMU ZARARI DETAYLARI
            st.info("Sekme 4: Kamu Zararı Detayları oluşturuluyor...")
            df_kz = df[df['_KamuZarariVar']].copy()
            if not df_kz.empty:
                kz_id_konu = df_kz[df_kz['Karar_Turu'] == 'İlk Derece Kararı']['Karar_Konusu'].value_counts().reset_index().rename(columns={'index':'Konu', 'Karar_Konusu':'Sayı'})
                kz_iade_konu = df_kz[df_kz['Karar_Turu'] == 'Yargılamanın İadesi sonucu verilen karar']['Karar_Konusu'].value_counts().reset_index().rename(columns={'index':'Konu', 'Karar_Konusu':'Sayı'})
                kz_id_konu.to_excel(writer, sheet_name='Kamu_Zarari_Detaylari', startrow=1, startcol=0, index=False); writer.sheets['Kamu_Zarari_Detaylari'].cell(1, 1).value = 'KZ Olan İlk Derece - Konular'
                kz_iade_konu.to_excel(writer, sheet_name='Kamu_Zarari_Detaylari', startrow=1, startcol=3, index=False); writer.sheets['Kamu_Zarari_Detaylari'].cell(1, 4).value = 'KZ Olan Y. İadesi - Konular'
            
            # SEKME 5: Y. İADESİ & ISRAR KARARLARI DETAYLARI
            st.info("Sekme 5: Y. İadesi & Israr Kararları Detayları oluşturuluyor...")
            df_iade = df[df['Karar_Turu'] == 'Yargılamanın İadesi sonucu verilen karar'].copy()
            df_israr = df[df['_IsrarVar']].copy()
            if not df_iade.empty:
                iade_konu = df_iade['Karar_Konusu'].value_counts().reset_index().rename(columns={'index':'Konu', 'Karar_Konusu':'Sayı'})
                iade_konu.to_excel(writer, sheet_name='Iade_ve_Israr_Detaylari', startrow=1, startcol=0, index=False); writer.sheets['Iade_ve_Israr_Detaylari'].cell(1, 1).value = 'Y. İadesi Karar Konuları'
            if not df_israr.empty:
                israr_konu = df_israr['Karar_Konusu'].value_counts().reset_index().rename(columns={'index':'Konu', 'Karar_Konusu':'Sayı'})
                israr_konu.to_excel(writer, sheet_name='Iade_ve_Israr_Detaylari', startrow=1, startcol=3, index=False); writer.sheets['Iade_ve_Israr_Detaylari'].cell(1, 4).value = 'Israr Edilen Kararlar - Konular'

        return output_buffer.getvalue()

    except FileNotFoundError:
        st.error(f"HATA: 'sorumlu.xlsx' dosyası bulunamadı. Lütfen GitHub deponuza 'app.py' ile aynı dizine yüklediğinizden emin olun.")
        return None
    except KeyError as e:
        st.error(f"HATA: 'sorumlu.xlsx' dosyasında beklenen bir sütun başlığı bulunamadı: {e}")
        return None
    except Exception as e:
        st.error(f"Rapor oluşturulurken beklenmedik bir hata oluştu: {e}")
        return None

# ==============================================================================
# BÖLÜM 3: GENEL UYGULAMA YAPISI VE AYARLAR
# ==============================================================================

try:
    api_key = st.secrets["GEMINI_API_KEY"]
    genai.configure(api_key=api_key)
    gemini_model = genai.GenerativeModel('gemini-1.5-pro-latest')
except Exception as e:
    st.error(f"Gemini API anahtarı yüklenirken bir hata oluştu: {e}")
    gemini_model = None

st.set_page_config(page_title="Hukuki Metin Analizi", layout="wide")

st.title("⚖️ Gelişmiş Hukuki Analiz Platformu")
st.markdown("Bu platform, iki ana araç sunar: **Bireysel Metin Analizi** ve **Toplu Veri Raporlama**.")
st.markdown("---")

@st.cache_resource
def load_all_models():
    file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "final_models_combined.pkl")
    try:
        with open(file_path, "rb") as f:
            return pickle.load(f)
    except FileNotFoundError:
        st.error(f"🚨 Tahmin modeli dosyası bulunamadı: '{file_path}'. Lütfen GitHub deponuza yükleyin.")
        return None

@st.cache_data
def load_excel_data():
    file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "SOMUT OLAY-PYHTON.xlsx")
    try:
        df = pd.read_excel(file_path)
        if 'GİRİŞ' not in df.columns or 'Tam Metin' not in df.columns:
            st.error(f"'{os.path.basename(file_path)}' dosyasında 'GİRİŞ' ve/veya 'Tam Metin' sütunları bulunamadı.")
            return None
        return df
    except FileNotFoundError:
        st.error(f"🚨 Özetleme için veri dosyası bulunamadı: '{file_path}'. Lütfen GitHub deponuza yükleyin.")
        return None

models_bundle = load_all_models()
df_data = load_excel_data()

def predict_case(text, law_vec, damage_vec, law_mdl, damage_mdl, classes):
    X_laws = law_vec.transform([text])
    law_prediction_vector = law_mdl.predict(X_laws)[0]
    predicted_laws = [classes[i] for i, val in enumerate(law_prediction_vector) if val == 1]
    X_damage = damage_vec.transform([text])
    damage_prediction_code = damage_mdl.predict(X_damage)[0]
    has_public_damage = "VAR" if damage_prediction_code == 1 else "YOK"
    return predicted_laws, has_public_damage

def find_full_text(df, input_text):
    if df is None or not input_text or not input_text.strip(): return None
    mask = df['GİRİŞ'].str.strip().str.startswith(input_text.strip(), na=False)
    return df.loc[mask, 'Tam Metin'].iloc[0] if mask.any() else None

def get_gemini_summary(text):
    if gemini_model is None: return "Gemini modeli yüklenemediği için özet oluşturulamadı."
    try:
        prompt = f"""Aşağıdaki hukuki metni analiz et ve ana konuyu, tarafların temel argümanlarını ve olayın sonucunu (eğer belirtilmişse) vurgulayan kısa ve anlaşılır bir özet çıkar. Özet, hukuki terimlerden arındırılmış ve herkesin anlayabileceği bir dilde olmalıdır. Metin: "{text}" Özet: """
        response = gemini_model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"Gemini özetleme sırasında bir hata oluştu: {e}"

# ==============================================================================
# BÖLÜM 4: KULLANICI ARAYÜZÜ (STREAMLIT UI)
# ==============================================================================

st.header("1. Bireysel Dava Metni Analizi")

if models_bundle is None or df_data is None:
    st.warning("Bireysel analiz aracı için gerekli model veya veri dosyaları yüklenemedi.")
else:
    law_model, damage_model, vectorizer_laws, vectorizer_damage, mlb_classes = (
        models_bundle['law_model'], models_bundle['damage_model'], 
        models_bundle['vectorizer_laws'], models_bundle['vectorizer_damage'], 
        models_bundle['mlb_classes']
    )
    col1, col2 = st.columns([2, 1])
    with col1:
        input_text = st.text_area("Analiz edilecek metnin başlangıcını girin:", height=250, placeholder="Örnek: Eşi çalışan personele aile yardımı ödeneği ödenmesi...")
        if st.button("🔍 Analiz Et", type="primary", use_container_width=True):
            if input_text.strip():
                with st.spinner("Analiz yapılıyor..."):
                    st.session_state.laws, st.session_state.damage = predict_case(input_text, vectorizer_laws, vectorizer_damage, law_model, damage_model, mlb_classes)
                    full_text = find_full_text(df_data, input_text)
                    st.session_state.summary = get_gemini_summary(full_text) if full_text else "Girdiğiniz metinle eşleşen bir 'Tam Metin' bulunamadı."
                    st.session_state.ran_prediction = True
            else:
                st.warning("Lütfen analiz için bir metin girin.")
    with col2:
        st.subheader("📊 Analiz Sonuçları")
        if st.session_state.get('ran_prediction', False):
            st.markdown("##### 📘 İlgili Kanunlar:")
            if st.session_state.laws:
                for k in st.session_state.laws: st.success(f"- {k}")
            else:
                st.info("İlişkili bir kanun bulunamadı.")
            
            st.markdown("---")
            st.markdown("##### 💸 Kamu Zararı Durumu:")
            st.error(f"**{st.session_state.damage}**") if st.session_state.damage == "VAR" else st.info(f"**{st.session_state.damage}**")
            
            st.markdown("---")
            st.markdown("##### 🤖 Gemini AI Metin Özeti:")
            with st.expander("Özeti Göster", expanded=True):
                st.info(st.session_state.summary)
        else:
            st.info("Sonuçları görmek için bir metin girip 'Analiz Et' butonuna tıklayın.")

st.markdown("\n\n---\n\n")

st.header("2. Toplu Veri Analizi ve Raporlama")
st.markdown("`sorumlu.xlsx` dosyasını kullanarak kapsamlı bir analiz yapar ve sonuçları grafiklerle zenginleştirilmiş yeni bir Excel dosyası olarak sunar.")

if st.button("📊 Kapsamlı Analiz Raporu Oluştur", use_container_width=True):
    with st.spinner("Rapor oluşturuluyor..."):
        script_dir = os.path.dirname(os.path.realpath(__file__))
        report_data = generate_excel_report(script_dir)
        if report_data:
            st.session_state.report_data = report_data
            st.success("✅ Rapor başarıyla oluşturuldu! Aşağıdaki butondan indirebilirsiniz.")

if 'report_data' in st.session_state and st.session_state.report_data:
    st.download_button(
        label="📥 Analiz Raporunu İndir (.xlsx)",
        data=st.session_state.report_data,
        file_name="Vaaaov_Analiz_Raporu.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
